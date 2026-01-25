"""
Scorecard Exporter
Export scorecard to Excel and PMML formats
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from sklearn.base import BaseEstimator, TransformerMixin
import io
import matplotlib.pyplot as plt
import seaborn as sns

class WOETransformer(BaseEstimator, TransformerMixin):
    """
    WOE transformation wrapper for PMML export.
    
    This class is picklable (unlike nested functions),
    which is required for sklearn2pmml serialization.
    """
    
    def __init__(self, binning_engine):
        """
        Initialize WOE transformer
        
        Args:
            binning_engine: BinningEngineWrapper with binned_results
        """
        self.binning_engine = binning_engine
    
    def fit(self, X, y=None):
        """
        Fit method (no-op, binning rules already exist)
        
        Args:
            X: Input data (not used, kept for sklearn compatibility)
            y: Target (not used)
        
        Returns:
            self
        """
        # Binning rules already exist in binning_engine
        # This is just for sklearn API compatibility
        return self
    
    def transform(self, X):
        """
        Transform raw data to WOE values
        
        Args:
            X: DataFrame with raw variable values
        
        Returns:
            DataFrame with WOE-transformed values
        """
        # Apply WOE transformation
        X_woe = self.binning_engine.binner.apply_bins(
            dataset=X,
            dict_woe=self.binning_engine.binned_results,
            is_df=False
        )
        
        return X_woe
    
    def get_params(self, deep=True):
        """Get parameters (for sklearn compatibility)"""
        return {'binning_engine': self.binning_engine}
    
    def set_params(self, **params):
        """Set parameters (for sklearn compatibility)"""
        if 'binning_engine' in params:
            self.binning_engine = params['binning_engine']
        return self

class ScorecardExporter:
    """Export scorecard results to various formats"""
    
    def __init__(self, model, train_scored, val_scored, metrics, correlation_df=None,X_train_woe=None):
        self.model = model
        self.train_scored = train_scored
        self.val_scored = val_scored
        self.metrics = metrics
        self.correlation_df = correlation_df
        self.X_train_woe = X_train_woe 
    
    def export_to_excel(self, filepath):
        """
        Export complete scorecard to Excel with multiple sheets
        
        Args:
            filepath: Path to save Excel file
        """
        print(f"[EXCEL] Creating workbook...")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Summary
        self._create_summary_sheet(wb)
        
        # Sheet 2: Scorecard
        self._create_scorecard_sheet(wb)
        
        # Sheet 3: Performance Metrics
        self._create_metrics_sheet(wb)
        
        # Sheet 4: Coefficient Analysis
        self._create_coefficient_sheet(wb)
        
        # Sheet 5: Correlation Matrix
        self._create_correlation_sheet(wb) 
        
        # Sheet 6: Plots (as images)
        self._create_plots_sheet(wb)
        
        # Save workbook
        wb.save(filepath)
        print(f"[EXCEL] Saved to {filepath}")
    
    def _create_summary_sheet(self, wb):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Credit Risk Scorecard - Summary"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Model info
        ws['A3'] = "Model Configuration"
        ws['A3'].font = Font(size=12, bold=True, color="2C3E50")
        
        config_data = [
            ["Parameter", "Value"],
            ["Base Score", self.model.score_params['base_score']],
            ["PDO", self.model.score_params['pdo']],
            ["Base Odds", self.model.score_params['base_odds']],
            ["Population Odds", f"{self.model.score_params['population_odds']:.4f}"],
            ["Number of Variables", len(self.model.scorecard_table['Variable'].unique()) - 1],  # -1 for BASE_SCORE
        ]
        
        for row_idx, row_data in enumerate(config_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:  # Header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
        
        # Performance summary
        ws['A11'] = "Performance Summary"
        ws['A11'].font = Font(size=12, bold=True, color="2C3E50")
        
        perf_data = [
            ["Metric", "Training", "Validation"],
            ["Sample Size", self.metrics['train']['count'], self.metrics['val']['count']],
            ["KS Statistic", f"{self.metrics['train']['ks']:.4f}", f"{self.metrics['val']['ks']:.4f}"],
            ["KS Score", f"{self.metrics['train']['ks_score']:.0f}", f"{self.metrics['val']['ks_score']:.0f}"],
            ["AUC", f"{self.metrics['train']['auc']:.4f}", f"{self.metrics['val']['auc']:.4f}"],
            ["Gini", f"{self.metrics['train']['gini']:.4f}", f"{self.metrics['val']['gini']:.4f}"],
        ]
        
        for row_idx, row_data in enumerate(perf_data, start=12):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 12:  # Header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="center")
        
        # Style
        self._apply_borders(ws, 'A4:B9')
        self._apply_borders(ws, 'A12:C17')
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def _create_scorecard_sheet(self, wb):
        """Create scorecard sheet"""
        ws = wb.create_sheet("Scorecard")
        
        # Title
        ws['A1'] = "Credit Scorecard - Point Allocation"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.merge_cells('A1:H1')
        
        # Get scorecard data (exclude BASE_SCORE)
        scorecard_df = self.model.scorecard_table[
            self.model.scorecard_table['Variable'] != 'BASE_SCORE'
        ].copy()
        
        # Select columns to display
        display_cols = ['Variable', 'Value', 'WOE', 'Coefficient', 'Points', 'Total', 'Target Rate']
        scorecard_display = scorecard_df[display_cols].copy()
        
        # Format
        scorecard_display['Total'] = scorecard_display['Total'].fillna(0).astype(int)
        scorecard_display['Target Rate'] = scorecard_display['Target Rate'].fillna(0)
        
        # Write header
        headers = ['Variable', 'Bin', 'WOE', 'Coefficient', 'Points', 'Count', 'Target Rate (%)']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, (_, row_data) in enumerate(scorecard_display.iterrows(), start=4):
            ws.cell(row=row_idx, column=1, value=row_data['Variable'])
            ws.cell(row=row_idx, column=2, value=str(row_data['Value']))
            ws.cell(row=row_idx, column=3, value=float(row_data['WOE'])).number_format = '0.0000'
            ws.cell(row=row_idx, column=4, value=float(row_data['Coefficient'])).number_format = '0.0000'
            
            # Points with conditional formatting
            points_cell = ws.cell(row=row_idx, column=5, value=int(row_data['Points']))
            if row_data['Points'] > 0:
                points_cell.fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
            elif row_data['Points'] < 0:
                points_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            
            ws.cell(row=row_idx, column=6, value=int(row_data['Total'])).number_format = '#,##0'
            ws.cell(row=row_idx, column=7, value=float(row_data['Target Rate'])).number_format = '0.00%'
            
            # Center alignment
            for col in range(3, 8):
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
        
        # Borders
        self._apply_borders(ws, f'A3:G{3 + len(scorecard_display)}')
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        
        # Add base score note
        base_row = self.model.scorecard_table[self.model.scorecard_table['Variable'] == 'BASE_SCORE']
        if not base_row.empty:
            base_points = int(base_row['Points'].values[0])
            note_row = 4 + len(scorecard_display) + 2
            ws[f'A{note_row}'] = "Base Score (Intercept):"
            ws[f'A{note_row}'].font = Font(bold=True)
            ws[f'E{note_row}'] = base_points
            ws[f'E{note_row}'].font = Font(bold=True)
            ws[f'E{note_row}'].alignment = Alignment(horizontal="center")
    
    def _create_metrics_sheet(self, wb):
        """Create detailed metrics sheet"""
        ws = wb.create_sheet("Performance Metrics")
        
        # Title
        ws['A1'] = "Detailed Performance Metrics"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Score distribution statistics
        ws['A3'] = "Score Distribution"
        ws['A3'].font = Font(size=12, bold=True, color="2C3E50")
        
        train_scores = self.train_scored['score']
        val_scores = self.val_scored['score']
        
        dist_data = [
            ["Statistic", "Training", "Validation"],
            ["Mean", f"{train_scores.mean():.2f}", f"{val_scores.mean():.2f}"],
            ["Median", f"{train_scores.median():.2f}", f"{val_scores.median():.2f}"],
            ["Std Dev", f"{train_scores.std():.2f}", f"{val_scores.std():.2f}"],
            ["Min", f"{train_scores.min():.2f}", f"{val_scores.min():.2f}"],
            ["Max", f"{train_scores.max():.2f}", f"{val_scores.max():.2f}"],
            ["25th Percentile", f"{train_scores.quantile(0.25):.2f}", f"{val_scores.quantile(0.25):.2f}"],
            ["75th Percentile", f"{train_scores.quantile(0.75):.2f}", f"{val_scores.quantile(0.75):.2f}"],
        ]
        
        for row_idx, row_data in enumerate(dist_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:  # Header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                else:
                    if col_idx > 1:
                        cell.alignment = Alignment(horizontal="center")
        
        # Classification metrics by score bands
        ws['A14'] = "Performance by Score Bands"
        ws['A14'].font = Font(size=12, bold=True, color="2C3E50")
        
        # Create score bands
        val_scored_copy = self.val_scored.copy()
        val_scored_copy['score_band'] = pd.cut(
            val_scored_copy['score'],
            bins=5,
            labels=['Very Low', 'Low', 'Medium', 'High', 'Very High']
        )
        
        band_stats = val_scored_copy.groupby('score_band').agg({
            'score': ['count', 'mean'],
            'target': ['sum', 'mean']
        }).round(4)
        
        band_stats.columns = ['Count', 'Avg Score', 'Defaults', 'Default Rate']
        band_stats.reset_index(inplace=True)
        
        # Write band stats
        band_headers = ['Score Band', 'Count', 'Avg Score', 'Defaults', 'Default Rate (%)']
        for col_idx, header in enumerate(band_headers, start=1):
            cell = ws.cell(row=15, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, (_, row_data) in enumerate(band_stats.iterrows(), start=16):
            ws.cell(row=row_idx, column=1, value=row_data['score_band'])
            ws.cell(row=row_idx, column=2, value=int(row_data['Count']))
            ws.cell(row=row_idx, column=3, value=float(row_data['Avg Score'])).number_format = '0.00'
            ws.cell(row=row_idx, column=4, value=int(row_data['Defaults']))
            ws.cell(row=row_idx, column=5, value=float(row_data['Default Rate'])).number_format = '0.00%'
            
            for col in range(2, 6):
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
        
        # Borders
        self._apply_borders(ws, 'A4:C11')
        self._apply_borders(ws, 'A15:E20')
        
        # Column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
    
    def _create_coefficient_sheet(self, wb):
        """Create coefficient analysis sheet"""
        ws = wb.create_sheet("Coefficient Analysis")
        
        # Title
        ws['A1'] = "Model Coefficients and Significance"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.merge_cells('A1:E1')
        
        # Get coefficient data
        scorecard_df = self.model.scorecard_table[
            self.model.scorecard_table['Variable'] != 'BASE_SCORE'
        ].copy()
        
        # Group by variable to get unique coefficients
        var_coefs = scorecard_df.groupby('Variable').agg({
            'Coefficient': 'first',
            'iv': 'sum'  # Sum IV contributions
        }).reset_index()
        var_coefs.columns = ['Variable', 'Coefficient', 'IV']
        var_coefs = var_coefs.sort_values('Coefficient', key=abs, ascending=False)
        
        # Write header
        headers = ['Variable', 'Coefficient', 'Information Value', 'Impact']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, (_, row_data) in enumerate(var_coefs.iterrows(), start=4):
            ws.cell(row=row_idx, column=1, value=row_data['Variable'])
            ws.cell(row=row_idx, column=2, value=float(row_data['Coefficient'])).number_format = '0.0000'
            ws.cell(row=row_idx, column=3, value=float(row_data['IV'])).number_format = '0.0000'
            
            # Impact assessment
            coef_abs = abs(row_data['Coefficient'])
            if coef_abs > 1.0:
                impact = "High"
                color = "E74C3C"
            elif coef_abs > 0.5:
                impact = "Medium"
                color = "F39C12"
            else:
                impact = "Low"
                color = "3498DB"
            
            impact_cell = ws.cell(row=row_idx, column=4, value=impact)
            impact_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            impact_cell.font = Font(color="FFFFFF", bold=True)
            impact_cell.alignment = Alignment(horizontal="center")
            
            # Center alignment
            for col in range(2, 4):
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
        
        # Add intercept
        intercept_row = 4 + len(var_coefs) + 2
        ws[f'A{intercept_row}'] = "Intercept (Base Score)"
        ws[f'A{intercept_row}'].font = Font(bold=True)
        ws[f'B{intercept_row}'] = float(self.model.classifier.intercept_[0])
        ws[f'B{intercept_row}'].number_format = '0.0000'
        ws[f'B{intercept_row}'].font = Font(bold=True)
        ws[f'B{intercept_row}'].alignment = Alignment(horizontal="center")
        
        # Borders
        self._apply_borders(ws, f'A3:D{3 + len(var_coefs)}')
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
    
    def _create_plots_sheet(self, wb):
        """Create sheet with embedded plots"""
        ws = wb.create_sheet("Diagnostic Plots")
        
        # Title
        ws['A1'] = "Model Diagnostic Visualizations"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.merge_cells('A1:H1')
        
        # Set matplotlib style
        plt.style.use('default')
        sns.set_palette(["#3498DB", "#E74C3C", "#27AE60", "#F39C12"])
        
        # Create plots and embed as images
        plots = [
            ("ROC Curve", self._create_roc_plot_image, 'A3'),
            ("KS Chart", self._create_ks_plot_image, 'A25'),
            ("Score Distribution", self._create_dist_plot_image, 'I3'),
            ("Calibration", self._create_cal_plot_image, 'I25'),
        ]
        
        for title, plot_func, anchor in plots:
            # Add title
            row = int(''.join(filter(str.isdigit, anchor)))
            ws[f'A{row}'] = title
            ws[f'A{row}'].font = Font(size=12, bold=True, color="2C3E50")
            
            # Create and embed plot
            img_stream = plot_func()
            img = XLImage(img_stream)
            img.width = 480
            img.height = 360
            ws.add_image(img, anchor)
        
        # Column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 2
    
    def _create_roc_plot_image(self):
        """Create ROC plot as image stream"""
        fig, ax = plt.subplots(figsize=(6, 4.5), dpi=100)
        
        from sklearn.metrics import roc_curve, auc
        
        # Training
        fpr_train, tpr_train, _ = roc_curve(self.train_scored['target'], -self.train_scored['score'])
        auc_train = auc(fpr_train, tpr_train)
        ax.plot(fpr_train, tpr_train, color='#3498DB', linewidth=2, label=f'Train AUC={auc_train:.3f}')
        
        # Validation
        fpr_val, tpr_val, _ = roc_curve(self.val_scored['target'], -self.val_scored['score'])
        auc_val = auc(fpr_val, tpr_val)
        ax.plot(fpr_val, tpr_val, color='#E74C3C', linewidth=2, label=f'Val AUC={auc_val:.3f}')
        
        # Diagonal
        ax.plot([0, 1], [0, 1], 'k--', linewidth=1, alpha=0.5)
        
        ax.set_xlabel('False Positive Rate', fontsize=10)
        ax.set_ylabel('True Positive Rate', fontsize=10)
        ax.set_title('ROC Curve', fontsize=11, fontweight='bold')
        ax.legend(loc='lower right', fontsize=9)
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        # Save to stream
        img_stream = io.BytesIO()
        fig.savefig(img_stream, format='png', dpi=100, bbox_inches='tight')
        img_stream.seek(0)
        plt.close(fig)
        
        return img_stream
    
    def _create_ks_plot_image(self):
        """Create KS plot as image stream"""
        fig, ax = plt.subplots(figsize=(6, 4.5), dpi=100)
        
        ks_stat, ks_score, ks_data = self.model.calculate_ks_statistic(self.val_scored)
        
        ax.plot(ks_data['score'], ks_data['cum_good_pct'], color='#27AE60', linewidth=2, label='Good')
        ax.plot(ks_data['score'], ks_data['cum_bad_pct'], color='#E74C3C', linewidth=2, label='Bad')
        
        ks_row = ks_data[ks_data['score'] == ks_score].iloc[0]
        ax.vlines(x=ks_score, ymin=ks_row['cum_bad_pct'], ymax=ks_row['cum_good_pct'],
                 colors='#3498DB', linewidth=2, linestyles='dashed', label=f'KS={ks_stat:.3f}')
        
        ax.set_xlabel('Credit Score', fontsize=10)
        ax.set_ylabel('Cumulative %', fontsize=10)
        ax.set_title('KS Chart', fontsize=11, fontweight='bold')
        ax.legend(loc='upper left', fontsize=9)
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        img_stream = io.BytesIO()
        fig.savefig(img_stream, format='png', dpi=100, bbox_inches='tight')
        img_stream.seek(0)
        plt.close(fig)
        
        return img_stream
    
    def _create_dist_plot_image(self):
        """Create distribution plot as image stream"""
        fig, ax = plt.subplots(figsize=(6, 4.5), dpi=100)
        
        good_scores = self.val_scored[self.val_scored['target'] == 0]['score']
        bad_scores = self.val_scored[self.val_scored['target'] == 1]['score']
        
        sns.kdeplot(data=good_scores, ax=ax, fill=True, color='#27AE60', alpha=0.5, linewidth=2, label='Good')
        sns.kdeplot(data=bad_scores, ax=ax, fill=True, color='#E74C3C', alpha=0.5, linewidth=2, label='Bad')
        
        ax.set_xlabel('Credit Score', fontsize=10)
        ax.set_ylabel('Density', fontsize=10)
        ax.set_title('Score Distribution', fontsize=11, fontweight='bold')
        ax.legend(loc='upper left', fontsize=9)
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        img_stream = io.BytesIO()
        fig.savefig(img_stream, format='png', dpi=100, bbox_inches='tight')
        img_stream.seek(0)
        plt.close(fig)
        
        return img_stream
    
    
    def _create_correlation_sheet(self, workbook):
        """Create correlation matrix sheet"""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        ws = workbook.create_sheet("Correlation Matrix")
        
        try:
            # Get selected variables
            selected_vars = [row['Variable'] for _, row in self.model.scorecard_table.iterrows() 
                            if row['Variable'] != 'BASE_SCORE']
            selected_vars = list(dict.fromkeys(selected_vars))  # Remove duplicates
            
            # Get correlation matrix from model
            # Note: You'll need to pass X_train_woe to exporter or store in model
            # For now, create a simple version
            
            # Title
            ws['A1'] = "Coefficient Correlation Matrix"
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells('A1:B3')
            
            # Subtitle
            ws['A4'] = "Correlation threshold: |r| > 0.4 highlighted"
            ws['A4'].font = Font(size=10, italic=True)
            ws.merge_cells('A4:B4')
            
            # Note: To get actual correlation values, you need to either:
            # 1. Pass correlation_df to the exporter
            # 2. Store it in the model
            # 3. Recalculate it here
            
            # For demonstration, let's assume correlation_df is available
            # You'll need to modify ScorecardExporter.__init__ to accept it
            
            if hasattr(self, 'correlation_df') and self.correlation_df is not None:
                corr_df = self.correlation_df
                
                start_row = 6
                start_col = 2
                
                # Write column headers
                for col_idx, var in enumerate(corr_df.columns):
                    cell = ws.cell(row=start_row, column=start_col + col_idx)
                    cell.value = var
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws.column_dimensions[get_column_letter(start_col + col_idx)].width = 15
                
                # Write row headers and data
                for row_idx, var in enumerate(corr_df.index):
                    # Row header
                    cell = ws.cell(row=start_row + 1 + row_idx, column=1)
                    cell.value = var
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    ws.column_dimensions['A'].width = 20
                    
                    # Correlation values
                    for col_idx, corr_var in enumerate(corr_df.columns):
                        corr_value = corr_df.loc[var, corr_var]
                        cell = ws.cell(row=start_row + 1 + row_idx, column=start_col + col_idx)
                        cell.value = corr_value
                        cell.number_format = '0.000'
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Color coding
                        if var == corr_var:
                            # Diagonal - gray
                            cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                            cell.font = Font(bold=True)
                        elif abs(corr_value) > 0.7:
                            # High correlation - red
                            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF")
                        elif abs(corr_value) > 0.4:
                            # Medium correlation - orange
                            cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                            cell.font = Font(color="FFFFFF")
                        
                        # Borders
                        cell.border = Border(
                            left=Side(style='thin', color='CCCCCC'),
                            right=Side(style='thin', color='CCCCCC'),
                            top=Side(style='thin', color='CCCCCC'),
                            bottom=Side(style='thin', color='CCCCCC')
                        )
            else:
                ws['A6'] = "Correlation matrix data not available"
                ws['A6'].font = Font(italic=True, color="999999")
            
        except Exception as e:
            print(f"[WARNING] Could not create correlation matrix sheet: {str(e)}")
            ws['A6'] = f"Error: {str(e)}"
    
    def _create_cal_plot_image(self):
        """Create calibration plot as image stream"""
        fig, ax = plt.subplots(figsize=(6, 4.5), dpi=100)
        
        # Simplified calibration chart
        for dataset, color, label in [(self.val_scored, '#3498DB', 'Validation')]:
            dataset_copy = dataset.copy()
            dataset_copy['score_bin'] = pd.qcut(dataset_copy['score'], q=10, duplicates='drop')
            
            agg = dataset_copy.groupby('score_bin').agg({
                'score': 'mean',
                'target': 'mean'
            }).reset_index()
            
            ax.scatter(agg['score'], agg['target'], color=color, s=50, alpha=0.7, label=label)
        
        ax.set_xlabel('Credit Score', fontsize=10)
        ax.set_ylabel('Observed Default Rate', fontsize=10)
        ax.set_title('Calibration Chart', fontsize=11, fontweight='bold')
        ax.legend(fontsize=9)
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        img_stream = io.BytesIO()
        fig.savefig(img_stream, format='png', dpi=100, bbox_inches='tight')
        img_stream.seek(0)
        plt.close(fig)
        
        return img_stream
    
    def _apply_borders(self, ws, range_string):
        """Apply borders to a range of cells"""
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for row in ws[range_string]:
            for cell in row:
                cell.border = thin_border
                
    
    def _export_binning_rules(self, filepath, selected_vars):
        """
        Export binning rules as JSON for production use
        
        Args:
            filepath: Path to save JSON file
            selected_vars: List of variables
        """
        try:
            import json
            
            print(f"[BINNING RULES] Exporting to {filepath}...")
            
            binning_rules = {}
            
            for var in selected_vars:
                if var not in self.model.binning_engine.binned_results:
                    continue
                
                var_bins = self.model.binning_engine.binned_results[var]
                var_type = self.model.binning_engine.variable_metrics[var]['type']
                
                rules = {
                    'variable': var,
                    'type': var_type,
                    'bins': []
                }
                
                for idx, row in var_bins.iterrows():
                    bin_rule = {
                        'bin_id': int(idx),
                        'woe': float(row['WOE'])
                    }
                    
                    if var_type == 'numeric':
                        # Numeric bins
                        start = row['interval_start_include']
                        end = row['interval_end_exclude']
                        
                        bin_rule['interval_start_include'] = None if pd.isna(start) else float(start)
                        bin_rule['interval_end_exclude'] = None if pd.isna(end) else float(end)
                        bin_rule['label'] = str(row['Value'])
                    else:
                        # Categorical bins
                        bin_rule['categories'] = row['Value'] if isinstance(row['Value'], list) else [row['Value']]
                        bin_rule['label'] = str(row['Value'])
                    
                    rules['bins'].append(bin_rule)
                
                binning_rules[var] = rules
            
            # Write to file
            with open(filepath, 'w') as f:
                json.dump(binning_rules, f, indent=2, default=str)
            
            print(f"[BINNING RULES] Exported {len(binning_rules)} variables to {filepath}")
            print(f"[BINNING RULES] Use this file to apply WOE transformation before calling PMML model")
            
        except Exception as e:
            print(f"[WARNING] Could not export binning rules: {str(e)}")

    
    def export_to_pmml(self, filepath, selected_vars):
        """
        Export model to PMML format (expects WOE-transformed input)
        
        Args:
            filepath: Path to save PMML file
            selected_vars: List of variable names used in model
            
        Note:
            The exported PMML expects WOE-transformed values as input,
            not raw values. You must apply WOE transformation before
            calling the model in production.
        """
        print(f"[PMML] Exporting to {filepath}")
        
        try:
            from sklearn2pmml import sklearn2pmml, PMMLPipeline
            import pandas as pd
            
            # Check if WOE data is available
            if self.X_train_woe is None:
                raise ValueError(
                    "WOE training data not available. "
                    "Cannot export PMML without WOE transformation data."
                )
            
            print(f"[PMML] Using WOE data shape: {self.X_train_woe.shape}")
            print(f"[PMML] Selected variables: {selected_vars}")
            
            # Get WOE column names (these are the model's input features)
            woe_columns = list(self.X_train_woe.columns)
            print(f"[PMML] WOE columns ({len(woe_columns)}): {woe_columns}")
            
            # Create a simple pipeline with ONLY the classifier
            # (No WOE transformation - that must be done externally)
            print("[PMML] Creating pipeline with classifier only...")
            pipeline = PMMLPipeline([
                ("classifier", self.model.classifier)
            ])
            
            # Fit with WOE-transformed data to set the schema
            print("[PMML] Setting up PMML schema with WOE data...")
            
            # Use WOE training data (already transformed)
            X_woe_sample = self.X_train_woe.head(100)
            y_sample = self.train_scored['target'].head(100)
            
            print(f"[PMML] Sample WOE data shape: {X_woe_sample.shape}")
            print(f"[PMML] Sample WOE columns: {list(X_woe_sample.columns)}")
            
            # "Fit" the pipeline (classifier is already trained, just setting schema)
            pipeline.fit(X_woe_sample, y_sample)
            
            # Export to PMML
            print(f"[PMML] Writing PMML file to {filepath}...")
            sklearn2pmml(
                pipeline,
                filepath,
                with_repr=True,
                debug=True
            )
            
            print(f"\n{'='*70}")
            print("PMML EXPORT SUCCESSFUL")
            print(f"{'='*70}")
            print(f"File: {filepath}")
            print(f"Input features: {len(woe_columns)} WOE-transformed variables")
            print(f"Output: Probability of default (class 1)")
            print(f"\n⚠️  IMPORTANT:")
            print(f"   This PMML model expects WOE-TRANSFORMED input values,")
            print(f"   NOT raw variable values!")
            print(f"\n   Production workflow:")
            print(f"   1. Apply WOE transformation to raw data")
            print(f"   2. Pass WOE values to this PMML model")
            print(f"   3. Model returns probability")
            print(f"{'='*70}\n")
            
            # Create a companion binning rules file
            self._export_binning_rules(filepath.replace('.pmml', '_binning_rules.json'), selected_vars)
            
            return True
            
        except ImportError:
            error_msg = (
                "sklearn2pmml not installed.\n\n"
                "To export PMML models, install with:\n"
                "pip install sklearn2pmml\n\n"
                "Falling back to manual PMML creation..."
            )
            print(f"[PMML] {error_msg}")
            self._create_manual_pmml(filepath, selected_vars)
            return False
            
        except Exception as e:
            print(f"[PMML ERROR] {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise
    
    def _create_manual_pmml(self, filepath, feature_names):
        """Create basic PMML file manually"""
        from xml.etree.ElementTree import Element, SubElement, ElementTree
        from xml.dom import minidom
        
        # Create PMML structure
        pmml = Element('PMML', {
            'version': '4.4',
            'xmlns': 'http://www.dmg.org/PMML-4_4'
        })
        
        # Header
        header = SubElement(pmml, 'Header', {
            'copyright': 'Generated by Credit Risk Workbench',
            'description': 'Credit Scorecard Model'
        })
        
        # Data dictionary
        data_dict = SubElement(pmml, 'DataDictionary', {
            'numberOfFields': str(len(feature_names) + 1)
        })
        
        for feature in feature_names:
            SubElement(data_dict, 'DataField', {
                'name': feature,
                'optype': 'continuous',
                'dataType': 'double'
            })
        
        SubElement(data_dict, 'DataField', {
            'name': 'target',
            'optype': 'categorical',
            'dataType': 'integer'
        })
        
        # Regression model
        model = SubElement(pmml, 'RegressionModel', {
            'modelName': 'CreditScorecard',
            'functionName': 'classification',
            'algorithmName': 'LogisticRegression'
        })
        
        # Mining schema
        mining_schema = SubElement(model, 'MiningSchema')
        for feature in feature_names:
            SubElement(mining_schema, 'MiningField', {'name': feature})
        SubElement(mining_schema, 'MiningField', {'name': 'target', 'usageType': 'target'})
        
        # Regression table
        reg_table = SubElement(model, 'RegressionTable', {
            'intercept': str(self.model.classifier.intercept_[0]),
            'targetCategory': '1'
        })
        
        for idx, feature in enumerate(feature_names):
            SubElement(reg_table, 'NumericPredictor', {
                'name': feature,
                'coefficient': str(self.model.classifier.coef_[0][idx])
            })
        
        # Write to file
        tree = ElementTree(pmml)
        xmlstr = minidom.parseString(ElementTree.tostring(pmml)).toprettyxml(indent="  ")
        
        with open(filepath, 'w') as f:
            f.write(xmlstr)
        
        print(f"[PMML] Manual PMML created at {filepath}")